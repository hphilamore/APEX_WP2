import pandas as pd
import numpy as np
from read_excel_data_funcs import *
import re
import scipy
from scipy.optimize import minimize
import matplotlib.pyplot as plt
import itertools
from sklearn.model_selection import train_test_split
from sklearn.linear_model import Ridge
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error
from sklearn.model_selection import cross_val_score
from xgboost import XGBRegressor
import pickle
from openpyxl import Workbook

# Create a new workbook
work_book = Workbook()

# Remove the default empty sheet
work_book.remove(work_book.active)

mfc_types_regex_mappings = {
        r"10\*10\s*AC": "10x10_AC",
        r"20\*30\s*AC": "20x30_AC",
        r"10\*10(?!\s*AC)": "10x10",  # match 10*10 NOT followed by AC
        r"20\*30(?!\s*AC)": "20x30"     # match 20*30 NOT followed by AC
    }

# Hyperparameter ranges to test for each model
ridge_params = [
                {"alpha": a}
                for a in np.logspace(-4,4,100)
            ]

xgb_params = []
for n_estimators in [30, 50, 100]:
    for max_depth in [2, 3, 4]:
        for learning_rate in [0.01,0.05,0.1]:

            xgb_params.append({
                "n_estimators": n_estimators,
                "max_depth": max_depth,
                "learning_rate": learning_rate,
                "random_state":42
            })

def write_dicts_to_sheet(wb, sheet_name, rows):
    """Write a list of dictionaries to a new Excel worksheet."""

    # # Create a new workbook
    # wb = Workbook()

    # # Remove the default empty sheet
    # wb.remove(wb.active)

    ws = wb.create_sheet(title=sheet_name)

    # Column headings
    ws.append(list(rows[0].keys()))

    # Data
    for row in rows:
        ws.append(list(row.values()))

    # # Save workbook to excel file
    # wb.save("model_performance.xlsx")

    # return ws

def filter_feature_data(data, sheet_name, col_name, year, resistance):

    # Load date-time index for selected mfc
    COD_date_time_index = data['COD date time index'][col_name].copy()

    # Load resistance data for selected mfc
    resistance_column = data['Resistance kOhms'][col_name].copy()

    # Load feature  for selected mfc
    feature = data[sheet_name][col_name].copy()

    # Filter column to include data from selected year and resistance only
    mask_cod = COD_date_time_index.year == year
    mask_resistance = [r == resistance for r in resistance_column]
    feature = [f for f, c, r in zip(feature, mask_cod, mask_resistance) if c and r]

    return feature


def all_combinations(lst):
    return [
        list(combo)
        for r in range(1, len(lst) + 1)
        for combo in itertools.combinations(lst, r)
    ]


def prepare_model_data(model_data, labels, features, test_size):

    # Create DataFrame from the dictionary to make it easier to remove NaN values
    df = pd.DataFrame({f: model_data[f] for f in labels + features})

    # Convert everything to numeric; invalid values become NaN
    df = df.apply(pd.to_numeric, errors='coerce')

    # Drop rows where any value is NaN
    df = df.dropna()

    # Create y and X data as numpy arrays
    y = df[labels].to_numpy()

    X = df[features].to_numpy()

    # Split into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=test_size, random_state=42
    )       

    return X_train, X_test, y_train, y_test 

def scale_model_features(X_train, X_test):
    # Scale all features to mean 0, std 1, to retain feature importance 
    # if scale_features:
    scaler = StandardScaler()
    X_train_model = scaler.fit_transform(X_train)
    X_test_model = scaler.transform(X_test)

    # else:
    #     scaler = None
    #     X_train_model = X_train
    #     X_test_model = X_test

    return X_train_model, X_test_model, scaler

def downsample_to_target(subset_data,
                         n_observations, 
                         target_data_points):
    """
    TODO: Different random seed for each config
    """
    random_number_generator = np.random.default_rng(42)
    selected_indices = random_number_generator.choice(
                                n_observations,
                                size=target_data_points,
                                replace=False # Repeat indices not allowed
                            )

    subset_data = {
        key: [values[i] for i in selected_indices]
        for key, values in subset_data.items()
    }

    return subset_data


def optimise_hyperparameters(X_train, 
                               X_test, 
                               y_train, 
                               y_test, 
                               configuration, 
                               results,
                               model_class, 
                               param_grid,
                               features,  
                               scaler, 
                               verbose=True):
    

    print("Evaluating models through hyperparameter grid search...")

    # # Scale all features to mean 0, std 1, to retain feature importance 
    # if scale_features:
    #     scaler = StandardScaler()
    #     X_train_model = scaler.fit_transform(X_train)
    #     X_test_model = scaler.transform(X_test)
    # else:
    #     scaler = None
    #     X_train_model = X_train
    #     X_test_model = X_test

    # Store scores to compare performance using different hyperparameter combinations 
    param_gridsearch_results = []

    # Tune hyperparameters for current configuration using grid search
    for params in param_grid:

        model = model_class(**params)

        print(f"Evaluating {model}")

        # Train the model
        model.fit(X_train, y_train)
   
        # Make predictions
        y_pred = model.predict(X_test)

        # Evaluate
        mse = mean_squared_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        mae = mean_absolute_error(y_test, y_pred)

        # Store result for current hyperparameter combination 
        param_result = {
            "mfc_types": configuration['subset_mfc_types'],
            "resistances": configuration['subset_resistances'],
            "years": configuration['subset_years'],
            "features": features,
            "r2": r2,
            "mse": mse,
            "mae": mae,
            "model": model.__class__.__name__,
            "parameters": params,
            # "alpha": alpha,
            # "coefficients": model.coef_.copy(),
            # "intercept": model.intercept_,
            "feature_scales": scaler if scaler==None else scaler.scale_.copy(),
            "n_samples": len(y_train) + len(y_pred)
            }
        
        # Store model-specific information
        if hasattr(model, "coef_"):
            # param_result["coefficients"] = model.coef_.copy()
            param_result["coefficients"] = model.coef_.tolist()
            param_result["intercept"] = float(model.intercept_[0])

        if hasattr(model, "feature_importances_"):
            # param_result["feature_importances"] = model.feature_importances_
            param_result["feature_importances"] = model.feature_importances_.tolist()

        if scaler is not None:
            param_result["feature_scales"] = scaler.scale_
        
        param_gridsearch_results.append(param_result)

    # Sort paramteter grid search results by R² descending
    param_gridsearch_sorted = sorted(param_gridsearch_results, key=lambda x: x["r2"], reverse=True)

    if verbose==True:
        print('Keep config', configuration['subset_mfc_types'],
                configuration['subset_resistances'],
                configuration['subset_years'])

    # # Store the best value for this configuration with hyperparameter tuning 
    # results.append(param_gridsearch_sorted[0])

    # Return the best value for this configuration with hyperparameter tuning 
    best_gridsearch_result = param_gridsearch_sorted[0]
    return best_gridsearch_result

def format_result(result, mfc_types_regex_mappings):
    """Format a model result for printing and Excel."""

    return {
        # "Model": result["model"],
        # "Window Length (h)": round(result["window_length"], 3),
        "N Samples": result["n_samples"],
        "MFC Types": ", ".join(
            mfc_types_regex_mappings[p]
            for p in result["mfc_types"]
        ),
        "Resistances (kOhm)": ", ".join(
            map(str, result["resistances"])
        ),
        "Years": ", ".join(
            map(str, result["years"])
        ),
        "R²": round(result["r2"], 3),
        "MAE": round(result["mae"], 3),
        "Model Parameters": ", ".join(
            f"{k}: {v:.3f}"
            for k, v in result["parameters"].items()
        ),
        "Terms": ", ".join(result["features"]),
        "Coefficients": ", ".join(
            map(str, result.get("coefficients", []))
        ),
        "Intercept": result.get("intercept", ""),
        "Feature Importances": ", ".join(
            map(str, result.get("feature_importances", []))
        ),
    }


def extract_best_configs(model_results, verbose=True):

    best_configs = []

    # for results in [results_ridge, results_xgboost]:
    for results in model_results:

        # print('Results', results)

        # Sort results by R² descending
        results_sorted = sorted(results, key=lambda x: x["r2"], reverse=True)

        # Store the result that gives the highest R2 value
        best_configs.append(results_sorted[0])

        # print('Num results', len(results_sorted))

        # Show top 3 configurations 
        top_3 = results_sorted[:3]
        if verbose == True:

            print("\nTop 3 configurations:\n")
            for i, res in enumerate(top_3, 1):

                print(f"--- Rank {i} ---")
                print()
                print("Model:", res["model"])
                print("MFC types:", [mfc_types_regex_mappings[p] for p in res["mfc_types"]])
                # print("MFC types:", res["mfc_types"])
                print("Resistances kOhm:", res["resistances"])
                print("Years:", res["years"])
                print("R²:", round(res["r2"], 3))
                print("MSE:", round(res["mse"], 3))
                print("MAE:", round(res["mae"], 3))
                # print("Alpha:", res["alpha"]),
                print("Model Parameters:", ", ".join(f"{k}: {v:.3f}" for k, v in res["parameters"].items())),
                print("N Samples:", res["n_samples"])
                # print("Terms:", "v_peak, p_peak, energy, resistance")
                print("Terms:", res["features"])
                if "coefficients" in res:
                    print("Coefficients:", res["coefficients"])
                if "intercept" in res:
                    print("Intercept:", res["intercept"])
                if "feature_importances" in res:
                    print("Feature Importances:", res["feature_importances"])
                # print("Coefficients:", [float(round(r, 3)) for r in res["coefficients"]])
                # print("Intercept:", res["intercept"])
                print()
    return best_configs


def format_subset_result(result):
    subset = result["subset"]

    return {
        "MFC Types": ", ".join(subset["subset_mfc_types"]),
        "Resistances (kOhm)": ", ".join(map(str, subset["subset_resistances"])),
        "Years": ", ".join(map(str, subset["subset_years"])),
        "N Data Points": result["n_data_points"],
        "Included": result["included"],
        "Reason": result["reason"],
    }


def compare_input_data_configurations(features, 
                                      labels,
                                   mfc_types_all, # all mfc types to test in configurations 
                                   resistances_all, # all resistances to test in configurations 
                                   years_all, # all years to test in configurations 
                                #    test_size=0.2, 
                                wb,
                                    models = [Ridge, XGBRegressor],
                                    model_params = [ridge_params, xgb_params],
                                    scale_features=[True, False],
                                #    min_data_points=25,
                                target_data_points = 25,
                                test_combinations = True,
                                   verbose=True,
                                   window_length=None):
    
    # Summary of which subsets of input data were compared
    subset_summary = []

    # Results of data combinations that give best performance 
    model_results = [[] for m in models] 
    
    # Get stored feature data
    with open('mfc_features.pkl', 'rb') as f:
        mfc_features = pickle.load(f)

    # print(type(mfc_features))
    # for i in list(mfc_features)[:2]:
    #     print(i)
    #     print(mfc_features[i])

    # Get mfc names from the column headings for the first feature in the imported data
    first_feature = next(iter(mfc_features))
    mfc_names = mfc_features[first_feature].keys()

    # Test every possible combination of MFCs/resistances/years
    if test_combinations:
        subsets = itertools.product(
            all_combinations(mfc_types_all),
            all_combinations(resistances_all),
            all_combinations(years_all)
        )

    # Test only single MFC × resistance × year combinations
    else:
        subsets = itertools.product(
            mfc_types_all,
            resistances_all,
            years_all
        )

    # Find model solution for each subset
    for subset_mfc_types, subset_resistances, subset_years in subsets:

        # If using single MFC × resistance × year combinations, convert them to lists
        if not test_combinations:
            subset_mfc_types = [subset_mfc_types]
            subset_resistances = [subset_resistances]
            subset_years = [subset_years]

        subset = {
            'subset_mfc_types': subset_mfc_types,
            'subset_resistances': subset_resistances,
            'subset_years': subset_years
        }

        print('Subset:', subset)

        # Flag to determine whether results of this subset are saved
        contains_combination_with_zero_data = False

        # Copy feature names to new empty data structure to hold aggregated filtered feature data
        subset_data = {k : [] for k in mfc_features}

        # Skip MFC columns any that aren't in this subset
        for name in mfc_names:
            if not any(re.search(i, name) for i in subset['subset_mfc_types']):
                continue

            # Filter feature data to include only years and resistances in current subset 
            for year in subset['subset_years']:
                for resistance in subset['subset_resistances']:
                    for feature in subset_data:
                        filtered_feature = filter_feature_data(mfc_features, 
                                                                feature, 
                                                                name, 
                                                                year, 
                                                                resistance)

                        # Detect MFC/resistance/year combinations with zero data points 
                        if len(filtered_feature) == 0:
                            contains_combination_with_zero_data = True
        
                        # Extend feature data for this subset with filtered data
                        subset_data[feature].extend(filtered_feature)

        # Remove rows containig NaNs from feature data for this subset 
        df_subset = pd.DataFrame(subset_data)

        # Number of observations in this subset
        for key in list(subset_data)[:1]:
            n_observations_original = len(subset_data[key])
            # print(n_observations_original)

        # Drop rows where columns (excluding date-time and list data columns) contain NaN
        df_subset = df_subset.dropna(subset=df_subset.columns[1:-2])
        n_observations = len(df_subset)
        # print(n_observations)

        # Convert back to dictionary of lists
        subset_data = df_subset.to_dict(orient='list')

        rejection_reason = ""
        if contains_combination_with_zero_data == True:
            rejection_reason += "Missing MFC/resistance/year combination"

        # Exclude subsets containing MFC/resistance/year combinations with zero data points
        if n_observations < target_data_points:
            rejection_reason += f", Too few data points {n_observations}, \
            threshold={target_data_points}"

        # Exclude subsets containing too few observations from comparison
        if (contains_combination_with_zero_data == True or 
            n_observations < target_data_points):
            subset_summary.append({'subset': subset,
                                    'n_data_points': n_observations,
                                    'included': False,
                                    'reason': rejection_reason
                                    })
        
        #  Prepare subset feature data for modelling 
        else: 
            # Randomly downsample larger subset data sets to target data points 
            subset_data = downsample_to_target(subset_data, 
                                               n_observations, 
                                               target_data_points)

            subset_summary.append({
                            'subset': subset,
                            'n_data_points': n_observations,
                            'included': True,
                            'reason': ''
                        })

            # Create train and test data
            X_train, X_test, y_train, y_test = prepare_model_data(
                        subset_data, 
                        labels=['COD'],
                        features=features,
                        test_size=0.2
                    ) 

            for model_class, params, scale, results in zip(models, 
                                                           model_params, 
                                                           scale_features, 
                                                           model_results):

                # Scale model features 
                if scale:
                    X_train, X_test, scaler = scale_model_features(X_train, X_test)
                else:
                    scaler = None

                # Tune hyperparameters and get model result
                result = optimise_hyperparameters(
                            X_train, 
                            X_test, 
                            y_train, 
                            y_test, 
                            subset, 
                            results=results, 
                            model_class=model_class, 
                            param_grid=params,
                            features=features,
                            scaler=scaler,
                            verbose=verbose)    

                # Store result for this input data subset and window length
                results.append(result)

                format_result(result, mfc_types_regex_mappings)

                           
    # Store which subsets were included/excluded and reason as excel sheet
    formatted_subsets = [format_subset_result(result) for result in subset_summary]
    sheet_name = f"Subset Summary, target data points {target_data_points}"
    write_dicts_to_sheet(wb, sheet_name, formatted_subsets)

    # Store the results for each model for this window size 
    for results, model in zip(model_results, models):

        # # Array to store  R2 values for each window size for plotting 
        # r2s = []

        # Get model name
        model_name = results[0]["model"]

        # Format results as list of dictionaries, showing best result for each subset tested
        formatted_results = [format_result(r, mfc_types_regex_mappings) for r in results]

        # Create Excel worksheet with model name as the worksheet name
        write_dicts_to_sheet(
            wb,
            model_name + f"Window length {window_length}",
            formatted_results
        )

        # Print results
        for row in formatted_results:
            print(f"Window length {window_length}")
            for key, value in row.items():
                print(f"{key}: {value}")
            print()

    # Save workbook to excel file
    wb.save("model_performance.xlsx")
    


    # if window_length is not None:
    #     sheet_name = f"Subset Summary {window_length:.3g}h"
    # else:
    #     sheet_name = "Subset Summary"

    

    # # Extract the best configurations for each model
    # best_configs = extract_best_configs(model_results, verbose=verbose)
    # return best_configs

if __name__ == '__main__':

    print("comparing")

    compare_input_data_configurations(
        features=[
            'Vpeak mV', 
            'Ppeak W', 
            'Energy J', 
            'Resistance kOhms', 
            # 'Vfinal mV', 
            # 'Pfinal W'
            ],
        labels=['COD'],
        mfc_types_all = [
            r"10\*10\s*AC",     # Carbon veil + activated carbon
            r"20\*30\s*AC",
            r"10\*10(?!\s*AC)", # Carbon veil
            r"20\*30(?!\s*AC)"
            ],
        resistances_all = [0.1, 1, 3],
        # resistances_all = [1],
        years_all = [2024, 2025],
        wb=work_book,
        # years_all = [2024],#, 2025],
        models = [Ridge, XGBRegressor],
        target_data_points = 25,
        # models = [Ridge],
        model_params= [ridge_params, xgb_params],
        # test_combinations=False,
        scale_features=[True, False],
        verbose=False
    )







